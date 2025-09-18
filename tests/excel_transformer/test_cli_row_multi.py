import sys
from pathlib import Path

import openpyxl

# Ensure repo root is on sys.path so `import src.excel_transformer.cli` works
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.excel_transformer.cli import main as cli_main  # noqa: E402


def _write_sample_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["原始记录", "计分", "原始记录-问题", "问题的积分", "标准回答", "猜测回答"])
    ws.append(["记录A", 1, "问题A", 5, "标准答A", "猜测答A"])
    ws.append(["记录B", 2, "问题B", 8, "标准答B", "猜测答B"])
    ws.append(["记录C", 3, "问题C", 2, "标准答C", "猜测答C"])
    wb.save(str(path))


def _write_many_rows_xlsx(path: Path, n: int = 16) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["原始记录", "计分", "原始记录-问题", "问题的积分", "标准回答", "猜测回答"])
    for i in range(n):
        idx = i + 1
        ws.append([f"记录{idx}", idx % 5, f"问题{idx}", (idx * 3) % 10, f"标准答{idx}", f"猜测答{idx}"])
    wb.save(str(path))


def test_cli_row_accepts_comma_separated_multi_rows_csv(tmp_path, monkeypatch):
    excel_path = tmp_path / "sample.xlsx"
    _write_sample_xlsx(excel_path)
    cfg_path = REPO_ROOT / "scripts" / "example_config.conf"

    monkeypatch.chdir(tmp_path)

    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "csv", "--row", "1,2"])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample.csv"
    assert out_csv.exists()
    content = out_csv.read_text(encoding="utf-8-sig")
    # header + 2 data rows
    assert content.count("\n") >= 2


def test_cli_row_accepts_range_multi_rows_xlsx(tmp_path, monkeypatch):
    excel_path = tmp_path / "sample.xlsx"
    _write_sample_xlsx(excel_path)
    cfg_path = REPO_ROOT / "scripts" / "example_config.conf"

    monkeypatch.chdir(tmp_path)

    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "xlsx", "--row", "1-2"])
    assert rc == 0

    out_xlsx = tmp_path / "output" / "sample.xlsx"
    assert out_xlsx.exists()
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active
    # header row + 2 data rows
    assert ws.max_row == 3


def test_cli_row_accepts_bracket_range_csv(tmp_path, monkeypatch):
    excel_path = tmp_path / "sample.xlsx"
    _write_sample_xlsx(excel_path)
    cfg_path = REPO_ROOT / "scripts" / "example_config.conf"

    monkeypatch.chdir(tmp_path)

    # Bracket syntax [1,2] meaning inclusive range 1..2
    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "csv", "--row", "[1,2]"])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample.csv"
    assert out_csv.exists()
    content = out_csv.read_text(encoding="utf-8-sig")
    # header + 2 data rows
    assert content.count("\n") >= 2


def test_cli_row_accepts_bracket_range_xlsx(tmp_path, monkeypatch):
    excel_path = tmp_path / "sample.xlsx"
    _write_sample_xlsx(excel_path)
    cfg_path = REPO_ROOT / "scripts" / "example_config.conf"

    monkeypatch.chdir(tmp_path)

    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "xlsx", "--row", "[1,2]"])
    assert rc == 0

    out_xlsx = tmp_path / "output" / "sample.xlsx"
    assert out_xlsx.exists()
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active
    # header row + 2 data rows
    assert ws.max_row == 3


def test_cli_row_accepts_mixed_numbers_and_bracket_range_csv(tmp_path, monkeypatch):
    excel_path = tmp_path / "sample_many.xlsx"
    _write_many_rows_xlsx(excel_path, n=16)
    cfg_path = REPO_ROOT / "scripts" / "example_config.conf"

    monkeypatch.chdir(tmp_path)

    expr = "1,4,7,[9,13],10"
    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "csv", "--row", expr])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample_many.csv"
    assert out_csv.exists()
    content = out_csv.read_text(encoding="utf-8-sig")
    # header + requested rows (1,4,7,9,10,11,12,13,10) -> at least 9 data lines
    assert content.strip().count("\n") >= 9


def test_cli_row_accepts_mixed_numbers_and_bracket_range_xlsx(tmp_path, monkeypatch):
    excel_path = tmp_path / "sample_many.xlsx"
    _write_many_rows_xlsx(excel_path, n=16)
    cfg_path = REPO_ROOT / "scripts" / "example_config.conf"

    monkeypatch.chdir(tmp_path)

    expr = "1,4,7,[9,13],10"
    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "xlsx", "--row", expr])
    assert rc == 0

    out_xlsx = tmp_path / "output" / "sample_many.xlsx"
    assert out_xlsx.exists()
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active
    # header + 9 data rows expected (allowing duplicates)
    assert ws.max_row >= 10
