import csv
import json
import sys
from pathlib import Path


# Ensure repo root is on sys.path so `import src.excel_transformer.cli` works
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.excel_transformer.cli import main as cli_main  # noqa: E402
from src.excel_transformer.config import load_config  # noqa: E402
from src.excel_transformer.transform import transform_rows  # noqa: E402


def _write_config(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


def test_grouped_csv_default_labels(tmp_path, monkeypatch):
    # Use provided sample workbook
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score",
          "原始记录-问题": "ask",
          "问题的积分": "ask_score",
          "标准回答": "answer-1",
          "猜测回答": "answer-2"
        }

        [out]
        { "原始记录": "record", "计分": "score" }
        { "原始记录-问题": "ask", "问题的积分": "ask_score" }
        { "原始记录": "record", "回答": { "name": "answer", "value": "answer-1", "ex": { "if": "score==2", "value": "answer-2" } } }
        """
    )
    cfg_path = tmp_path / "cfg.conf"
    _write_config(cfg_path, cfg_text)

    monkeypatch.chdir(tmp_path)

    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "csv", "--grouped", "--row", "1"])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample.csv"
    assert out_csv.exists()

    with out_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        row = next(reader)

    assert header == ["group1", "group2", "group3"]

    g1 = json.loads(row[0])
    g2 = json.loads(row[1])
    g3 = json.loads(row[2])

    # Row 1 in sample.xlsx: 记录A / 1 / 问题A / 5 / 标准答A / 猜测答A
    assert g1 == {"原始记录": "记录A", "计分": 1}
    assert g2 == {"原始记录-问题": "问题A", "问题的积分": 5}
    assert g3 == {"原始记录": "记录A", "回答": "标准答A"}


def test_grouped_csv_custom_labels(tmp_path, monkeypatch):
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score",
          "原始记录-问题": "ask",
          "问题的积分": "ask_score",
          "标准回答": "answer-1",
          "猜测回答": "answer-2"
        }

        [out]
        { "__label__": "input", "原始记录": "record", "计分": "score" }
        { "__label__": "check", "原始记录-问题": "ask", "问题的积分": "ask_score" }
        { "__label__": "answer", "原始记录": "record", "回答": { "name": "answer", "value": "answer-1", "ex": { "if": "score==2", "value": "answer-2" } } }
        """
    )
    cfg_path = tmp_path / "cfg.conf"
    _write_config(cfg_path, cfg_text)

    monkeypatch.chdir(tmp_path)

    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "csv", "--grouped", "--row", "1"])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample.csv"
    assert out_csv.exists()

    with out_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        row = next(reader)

    assert header == ["input", "check", "answer"]

    g1 = json.loads(row[0])
    g2 = json.loads(row[1])
    g3 = json.loads(row[2])

    assert g1 == {"原始记录": "记录A", "计分": 1}
    assert g2 == {"原始记录-问题": "问题A", "问题的积分": 5}
    assert g3 == {"原始记录": "记录A", "回答": "标准答A"}


def test_grouped_xlsx_custom_labels(tmp_path, monkeypatch):
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score",
          "原始记录-问题": "ask",
          "问题的积分": "ask_score",
          "标准回答": "answer-1",
          "猜测回答": "answer-2"
        }

        [out]
        { "__label__": "input", "原始记录": "record", "计分": "score" }
        { "__label__": "check", "原始记录-问题": "ask", "问题的积分": "ask_score" }
        { "__label__": "answer", "原始记录": "record", "回答": { "name": "answer", "value": "answer-1", "ex": { "if": "score==2", "value": "answer-2" } } }
        """
    )
    cfg_path = tmp_path / "cfg.conf"
    _write_config(cfg_path, cfg_text)

    monkeypatch.chdir(tmp_path)

    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "xlsx", "--grouped", "--row", "1"])
    assert rc == 0

    out_xlsx = tmp_path / "output" / "sample.xlsx"
    assert out_xlsx.exists()

    import openpyxl

    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ["input", "check", "answer"]

    row = [c.value for c in ws[2]]
    g1 = json.loads(row[0])
    g2 = json.loads(row[1])
    g3 = json.loads(row[2])

    assert g1 == {"原始记录": "记录A", "计分": 1}
    assert g2 == {"原始记录-问题": "问题A", "问题的积分": 5}
    assert g3 == {"原始记录": "记录A", "回答": "标准答A"}


def test_grouped_csv_multiple_rows_pretty_json(tmp_path, monkeypatch):
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score",
          "原始记录-问题": "ask",
          "问题的积分": "ask_score",
          "标准回答": "answer-1",
          "猜测回答": "answer-2"
        }

        [out]
        { "__label__": "input", "原始记录": "record", "计分": "score" }
        { "__label__": "check", "原始记录-问题": "ask", "问题的积分": "ask_score" }
        { "__label__": "answer", "原始记录": "record", "回答": { "name": "answer", "value": "answer-1", "ex": { "if": "score==2", "value": "answer-2" } } }
        """
    )
    cfg_path = tmp_path / "cfg.conf"
    _write_config(cfg_path, cfg_text)

    monkeypatch.chdir(tmp_path)

    rc = cli_main([
        str(excel_path), "-c", str(cfg_path), "-f", "csv", "--grouped", "--rows", "1,2", "--pretty-json",
    ])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample.csv"
    assert out_csv.exists()

    rows = []
    with out_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        for r in reader:
            rows.append(r)

    assert header == ["input", "check", "answer"]
    assert len(rows) == 2

    # Pretty JSON should contain newlines
    assert "\n" in rows[0][0] and "\n" in rows[1][2]

    g1_r1 = json.loads(rows[0][0])
    g2_r1 = json.loads(rows[0][1])
    g3_r1 = json.loads(rows[0][2])
    g1_r2 = json.loads(rows[1][0])
    g2_r2 = json.loads(rows[1][1])
    g3_r2 = json.loads(rows[1][2])

    # Row 1 expectations
    assert g1_r1 == {"原始记录": "记录A", "计分": 1}
    assert g2_r1 == {"原始记录-问题": "问题A", "问题的积分": 5}
    assert g3_r1 == {"原始记录": "记录A", "回答": "标准答A"}

    # Row 2 expectations (score==2 -> use 猜测答B)
    assert g1_r2 == {"原始记录": "记录B", "计分": 2}
    assert g2_r2 == {"原始记录-问题": "问题B", "问题的积分": 8}
    assert g3_r2 == {"原始记录": "记录B", "回答": "猜测答B"}


def test_grouped_xlsx_multiple_rows_pretty_json(tmp_path, monkeypatch):
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score",
          "原始记录-问题": "ask",
          "问题的积分": "ask_score",
          "标准回答": "answer-1",
          "猜测回答": "answer-2"
        }

        [out]
        { "__label__": "input", "原始记录": "record", "计分": "score" }
        { "__label__": "check", "原始记录-问题": "ask", "问题的积分": "ask_score" }
        { "__label__": "answer", "原始记录": "record", "回答": { "name": "answer", "value": "answer-1", "ex": { "if": "score==2", "value": "answer-2" } } }
        """
    )
    cfg_path = tmp_path / "cfg.conf"
    _write_config(cfg_path, cfg_text)

    monkeypatch.chdir(tmp_path)

    rc = cli_main([
        str(excel_path), "-c", str(cfg_path), "-f", "xlsx", "--grouped", "--rows", "1,2", "--pretty-json",
    ])
    assert rc == 0

    out_xlsx = tmp_path / "output" / "sample.xlsx"
    assert out_xlsx.exists()

    import openpyxl

    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ["input", "check", "answer"]

    # Two data rows
    row1 = [c.value for c in ws[2]]
    row2 = [c.value for c in ws[3]]

    # Pretty JSON should contain newlines
    assert "\n" in row1[0] and "\n" in row2[2]

    g1_r1 = json.loads(row1[0])
    g2_r1 = json.loads(row1[1])
    g3_r1 = json.loads(row1[2])
    g1_r2 = json.loads(row2[0])
    g2_r2 = json.loads(row2[1])
    g3_r2 = json.loads(row2[2])

    assert g1_r1 == {"原始记录": "记录A", "计分": 1}
    assert g2_r1 == {"原始记录-问题": "问题A", "问题的积分": 5}
    assert g3_r1 == {"原始记录": "记录A", "回答": "标准答A"}
    assert g1_r2 == {"原始记录": "记录B", "计分": 2}
    assert g2_r2 == {"原始记录-问题": "问题B", "问题的积分": 8}
    assert g3_r2 == {"原始记录": "记录B", "回答": "猜测答B"}


def test_grouped_csv_label_line_syntax_headers(tmp_path, monkeypatch):
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score",
          "原始记录-问题": "ask",
          "问题的积分": "ask_score",
          "标准回答": "answer-1",
          "猜测回答": "answer-2"
        }

        [out]
        输入区:
        { "原始记录": "record", "计分": "score" }
        校验区:
        { "原始记录-问题": "ask", "问题的积分": "ask_score" }
        回答区:
        { "原始记录": "record", "回答": { "name": "answer", "value": "answer-1", "ex": { "if": "score==2", "value": "answer-2" } } }
        """
    )
    cfg_path = tmp_path / "cfg_label_lines.conf"
    _write_config(cfg_path, cfg_text)

    monkeypatch.chdir(tmp_path)
    rc = cli_main([str(excel_path), "-c", str(cfg_path), "-f", "csv", "--grouped", "--row", "1"])
    assert rc == 0

    out_csv = tmp_path / "output" / "sample.csv"
    with out_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)

    assert header == ["输入区", "校验区", "回答区"]


def test_terminal_output_includes_label_value(tmp_path):
    excel_path = REPO_ROOT / "tests" / "data" / "sample.xlsx"

    cfg_text = (
        """
        [map]
        {
          "原始记录": "record",
          "计分": "score"
        }

        [out]
        输入区:
        { "__label__": "输入区", "原始记录": "record", "计分": "score" }
        """
    )
    cfg_path = tmp_path / "cfg_terminal_label.conf"
    _write_config(cfg_path, cfg_text)

    cfg = load_config(str(cfg_path))
    rows = transform_rows(
        excel_path=str(excel_path),
        display_to_internal=cfg.display_to_internal,
        out_groups=cfg.out_groups,
        header_row=1,
        row_numbers=[1],
    )
    # Only one group -> single record with __label__ backfilled
    assert rows[0]["__label__"] == "输入区"
