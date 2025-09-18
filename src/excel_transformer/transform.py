from __future__ import annotations

import json
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from contextlib import contextmanager
import warnings

from openpyxl import load_workbook, Workbook


def _build_internal_to_display(mapper: Mapping[str, str]) -> Dict[str, str]:
    return {v: k for k, v in mapper.items()}


def _coerce_literal(val: Any) -> Any:
    # Try to coerce strings that look like numbers
    if isinstance(val, str):
        s = val.strip()
        if s.isdigit():
            try:
                return int(s)
            except Exception:
                return val
        try:
            return float(s)
        except Exception:
            return val
    return val


def _eval_condition(expr: str, context: Mapping[str, Any]) -> bool:
    # Very small evaluator supporting comparisons like a==b, a!=b, a>b, etc.
    ops = ["==", "!=", ">=", "<=", ">", "<"]
    for op in ops:
        if op in expr:
            left, right = expr.split(op, 1)
            left, right = left.strip(), right.strip()
            lv = context.get(left)
            rv: Any
            # Strip quotes around right if present
            if (right.startswith("'") and right.endswith("'")) or (
                right.startswith('"') and right.endswith('"')
            ):
                rv = right[1:-1]
            else:
                rv = _coerce_literal(right)
            lv = _coerce_literal(lv)
            try:
                if op == "==":
                    return lv == rv
                if op == "!=":
                    return lv != rv
                if op == ">=":
                    return lv >= rv  # type: ignore[operator]
                if op == "<=":
                    return lv <= rv  # type: ignore[operator]
                if op == ">":
                    return lv > rv  # type: ignore[operator]
                if op == "<":
                    return lv < rv  # type: ignore[operator]
            except Exception:
                return False
    # Fallback: unknown expression -> False
    return False


def _cell_value_by_display(headers: List[str], row_values: List[Any], display_name: str) -> Any:
    try:
        idx = headers.index(display_name)
    except ValueError:
        return ""
    if idx < 0 or idx >= len(row_values):
        return ""
    return row_values[idx]


def _value_by_internal(
    headers: List[str], row_values: List[Any], internal_to_display: Mapping[str, str], internal_key: str
) -> Any:
    display = internal_to_display.get(internal_key)
    if not display:
        return ""
    return _cell_value_by_display(headers, row_values, display)


def _resolve_field(
    field_key: str,
    spec: Any,
    headers: List[str],
    row_values: List[Any],
    internal_to_display: Mapping[str, str],
    context_by_internal: Mapping[str, Any],
) -> Tuple[str, Any]:
    """Resolve one field of an out-group into a key/value pair.

    Supports:
    - Direct mapping: spec is a string referencing an internal key -> fetch cell.
    - Rule mapping: spec is an object with optional keys {name, value, ex}.
    - Nested objects: spec is an object whose keys are not limited to {name, value, ex};
      in this case, recursively resolve each child to produce a nested dict. This supports
      multi-level nesting (情况三及更深层次)。
    """
    # Case 1: direct mapping (spec is internal key to fetch; output key is the field key)
    if isinstance(spec, str):
        output_key = field_key
        value = _value_by_internal(headers, row_values, internal_to_display, spec)
        return output_key, value

    if isinstance(spec, dict):
        # Distinguish between a rule object {name?, value?, ex?} and a nested object
        reserved = {"name", "value", "ex"}
        has_non_reserved_keys = any(k not in reserved for k in spec.keys())

        if has_non_reserved_keys:
            # Treat as nested object: recursively resolve each child
            nested: Dict[str, Any] = {}
            for child_key, child_spec in spec.items():
                ck, cv = _resolve_field(
                    field_key=str(child_key),
                    spec=child_spec,
                    headers=headers,
                    row_values=row_values,
                    internal_to_display=internal_to_display,
                    context_by_internal=context_by_internal,
                )
                nested[ck] = cv
            return str(field_key), nested

        # Otherwise, treat as rule mapping object
        name = spec.get("name")
        value_ref = spec.get("value")
        ex = spec.get("ex")
        # Evaluate ex condition to override value_ref
        if ex:
            # normalize to list of one rule or many
            rules = ex if isinstance(ex, list) else [ex]
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                condition = rule.get("if")
                if condition and _eval_condition(str(condition), context_by_internal):
                    # Prefer explicit 'value' in rule; else accept a key matching name
                    override = rule.get("value")
                    if not override and name:
                        override = rule.get(name)
                    if override is not None:
                        value_ref = override
                        break
        # Pull the value from the row by internal key reference
        # Output key should be the field key in the [out] group; `name` is a logical name, not the output label.
        output_key = str(field_key)
        if isinstance(value_ref, str):
            value = _value_by_internal(headers, row_values, internal_to_display, value_ref)
        else:
            value = value_ref
        return output_key, value

    # Fallback: emit as string
    return str(field_key), str(spec)


def transform_rows(
    excel_path: str,
    display_to_internal: Mapping[str, str],
    out_groups: Sequence[Mapping[str, Any]],
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    row_numbers: Optional[Sequence[int]] = None,  # 1-based, excluding header row
) -> List[Dict[str, Any]]:
    """Transform rows according to config into a list of output dicts.

    For each source row, each out-group produces one output record.
    """
    with _suppress_openpyxl_default_style_warning():
        wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read headers from header_row
    headers: List[str] = []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value is not None else "")

    internal_to_display = _build_internal_to_display(display_to_internal)

    results: List[Dict[str, Any]] = []

    # Build iterable of row indices to process (1-based including header in ws indexing)
    start_row_idx = header_row + 1
    end_row_idx = ws.max_row
    indices = range(start_row_idx, end_row_idx + 1)
    if row_numbers:
        # Convert to worksheet row indexes
        indices = [header_row + n for n in row_numbers]

    for r in indices:
        row_cells = ws[r]
        row_values = [c.value for c in row_cells]
        # Context by internal keys for condition evaluation
        context = {internal: _value_by_internal(headers, row_values, internal_to_display, internal)
                   for internal in internal_to_display.keys()}

        for group in out_groups:
            out: Dict[str, Any] = {}
            for key, spec in group.items():
                out_key, out_val = _resolve_field(
                    field_key=str(key),
                    spec=spec,
                    headers=headers,
                    row_values=row_values,
                    internal_to_display=internal_to_display,
                    context_by_internal=context,
                )
                out[out_key] = out_val
            results.append(out)

    return results


def transform_rows_grouped(
    excel_path: str,
    display_to_internal: Mapping[str, str],
    out_groups: Sequence[Mapping[str, Any]],
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    row_numbers: Optional[Sequence[int]] = None,  # 1-based, excluding header row
    *,
    group_label_key: str = "__label__",
) -> List[Dict[str, Any]]:
    """Transform rows into one record per source row, grouping each [out] object
    under a column named by a group label.

    - Group label: taken from a reserved key (default: "__label__") inside each
      [out] object. If absent, auto-named as "group1", "group2", ... by order.
    - Group value: the resolved object built from the remaining keys in that [out]
      object (excluding the reserved label key).
    - Result: each row yields a single dict like {label1: {...}, label2: {...}}.
    """
    with _suppress_openpyxl_default_style_warning():
        wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read headers from header_row
    headers: List[str] = []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value is not None else "")

    internal_to_display = _build_internal_to_display(display_to_internal)

    results: List[Dict[str, Any]] = []

    # Build iterable of row indices to process (1-based including header in ws indexing)
    start_row_idx = header_row + 1
    end_row_idx = ws.max_row
    indices = range(start_row_idx, end_row_idx + 1)
    if row_numbers:
        # Convert to worksheet row indexes
        indices = [header_row + n for n in row_numbers]

    for r in indices:
        row_cells = ws[r]
        row_values = [c.value for c in row_cells]
        # Context by internal keys for condition evaluation
        context = {
            internal: _value_by_internal(headers, row_values, internal_to_display, internal)
            for internal in internal_to_display.keys()
        }

        grouped: Dict[str, Any] = {}
        for idx, group in enumerate(out_groups):
            label = None
            if isinstance(group, dict) and group_label_key in group and isinstance(group[group_label_key], str):
                label = str(group[group_label_key])
            if not label:
                label = f"group{idx + 1}"

            obj: Dict[str, Any] = {}
            if isinstance(group, Mapping):
                for key, spec in group.items():
                    if str(key) == group_label_key:
                        continue  # skip label from object content
                    out_key, out_val = _resolve_field(
                        field_key=str(key),
                        spec=spec,
                        headers=headers,
                        row_values=row_values,
                        internal_to_display=internal_to_display,
                        context_by_internal=context,
                    )
                    obj[out_key] = out_val
            grouped[label] = obj

        results.append(grouped)

    return results


def write_csv(path: str, rows: List[Dict[str, Any]], *, compact_json: bool = True) -> None:
    import csv

    # Helper: convert any complex value to JSON string (pretty or compact)
    def _to_json_str(v: Any) -> str:
        if v is None:
            return ""
        # For containers and non-primitive types, dump as pretty JSON
        if isinstance(v, (dict, list, tuple, set)):
            try:
                to_dump = list(v) if isinstance(v, set) else v
                if compact_json:
                    return json.dumps(to_dump, ensure_ascii=False, separators=(",", ":"))
                return json.dumps(to_dump, ensure_ascii=False, indent=2)
            except Exception:
                return json.dumps(str(v), ensure_ascii=False)
        # Primitives: keep as-is but ensure string type for CSV safety
        if isinstance(v, (int, float, bool, str)):
            return str(v)
        # Fallback to string
        return str(v)

    # Build header order by first-seen key order across rows, preserving
    # the group/field order from config (JSON object order is preserved).
    fieldnames: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in fieldnames:
                fieldnames.append(k)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames,
            quoting=csv.QUOTE_ALL,  # quote all to avoid delimiter/newline induced misalignment
            lineterminator="\n",
        )
        writer.writeheader()
        for r in rows:
            serialized = {k: _to_json_str(r.get(k, "")) for k in fieldnames}
            writer.writerow(serialized)


def write_xlsx(path: str, rows: List[Dict[str, Any]], *, compact_json: bool = True) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "output"
    # Build header order by first-seen key order across rows to reflect
    # the [out] group/field sequence in config.
    headers: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in headers:
                headers.append(k)
    ws.append(headers)
    # Helper: ensure cell-friendly values; complex values -> pretty JSON string
    def _cell_safe(v: Any) -> Any:
        if v is None:
            return ""
        # Allow simple scalars
        if isinstance(v, (int, float, bool, str)):
            return v
        # Convert containers and other types to pretty JSON string
        try:
            to_dump = list(v) if isinstance(v, set) else v
            if compact_json:
                return json.dumps(to_dump, ensure_ascii=False, separators=(",", ":"))
            return json.dumps(to_dump, ensure_ascii=False, indent=2)
        except Exception:
            return str(v)

    for r in rows:
        ws.append([_cell_safe(r.get(h, "")) for h in headers])
    with _suppress_openpyxl_default_style_warning():
        wb.save(path)


def print_terminal(rows: List[Dict[str, Any]], pretty: bool = False) -> None:
    for r in rows:
        if pretty:
            print(json.dumps(r, ensure_ascii=False, indent=2))
        else:
            print(json.dumps(r, ensure_ascii=False))


@contextmanager
def _suppress_openpyxl_default_style_warning():
    """Suppress the common openpyxl warning when workbooks lack default style.

    Message: "Workbook contains no default style, apply openpyxl's default"
    Scope-limited to openpyxl.styles.stylesheet and only this specific message.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
            module=r"openpyxl\.styles\.stylesheet",
        )
        yield
